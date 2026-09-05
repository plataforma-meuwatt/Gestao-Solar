"""Confere, ponta a ponta, que o cliente baixa a planilha da usina pelo caminho do portal.

Não é teste automatizado, e é de propósito: bate no meuWatt de verdade, com a credencial
de serviço de verdade, e **abre os arquivos que saíram** para contar abas, linhas e colunas.
É a única prova que existe de que "Baixar dados" funciona — todo o resto (as medições da
mw-api direto, os testes com `respx`) prova pedaços.

A distinção que motiva o script: as medições anteriores desta feature foram feitas contra
a mw-api, com o PAT na mão. Isso não prova o caminho INTEIRO, que é onde mora o risco:

    `_usina_no_escopo` (usina_id nosso → PlantLink) → slug do banco, nunca do cliente
      → credencial de serviço → `httpx.stream` → `StreamingResponse` → bytes no navegador

Cada seta acima já quebrou em algum ponto deste repositório. E o arquivo que chega ao
cliente só é bom se ABRIR: um XLSX truncado pela recodificação do corpo, ou com a aba
Leia-me faltando, sai daqui com 200 e chega ao Excel como "arquivo corrompido".

**Baixa DOIS arquivos, não um.** O pesado (31 dias × 5 min, os quatro blocos, inversores em
lista) prova que o pior pedido possível atravessa. O do diretor (mesmo período, só o bloco
de inversores, agrupado POR SKID, com apenas 2 dos 5 skids marcados) prova a única coisa que
o pesado não prova: que a SELEÇÃO PARCIAL atravessou o BFF e **mudou a soma**. É a capacidade
que o dono recusou ver escondida atrás de um "avançado", e a que mais facilmente se perderia
em silêncio — porque um arquivo com a usina inteira também abre, também tem as abas certas e
também parece certo.

    python scripts/conferir_exportacao.py
    python scripts/conferir_exportacao.py --url https://gestao-solar-production.up.railway.app
    python scripts/conferir_exportacao.py --usina 4 --dias 31 --passo 5m --guardar
    python scripts/conferir_exportacao.py --provar     # não fala com ninguém: ver abaixo

Sem `--url`, sobe o BFF local num uvicorn próprio, em porta livre, e fala com ele por HTTP
de verdade. Não é `ASGITransport`: o que se quer provar é o `StreamingResponse` saindo por
um socket, com os cabeçalhos que o navegador vai ler — `Content-Disposition`,
`Content-Length` (ou a sua ausência declarada) e `Content-Type`. E é por rodar num processo
separado que dá para MEDIR a memória do worker enquanto os megabytes atravessam, que é a
única prova de que o fluxo é fluxo e não um buffer com outro nome.

`--provar` é o contrário: não fala com ninguém e alimenta as MESMAS conferências com dados
deliberadamente estragados (o slug do meuWatt vazando nas opções, a aba Leia-me faltando,
cinco colunas de skid onde foram pedidos dois, os dois 404 do escopo com frases diferentes),
exigindo que cada uma REPROVE. Um gate que nunca se viu reprovar é decoração; este se vê.

O `openpyxl` e o `psutil` não estão no `requirements.txt` porque o BFF não lê planilha nem
se mede: ele repassa bytes. Aqui são ferramenta de conferência. `pip install openpyxl psutil`
se faltarem — sem `psutil` a medida de memória é pulada em voz alta, o resto roda.

Sai com 0 só quando TODAS as conferências passam. Qualquer falha imprime o que se esperava,
o que veio, e sai com 1 — para poder ser encadeado num gate sem alguém ter de ler a saída.
"""

from __future__ import annotations

import argparse
import io
import json
import os
import socket
import subprocess
import sys
import tempfile
import threading
import time
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Callable

import httpx

# A saída tem acento e vai para um console cp1252 no Windows. Sem isto, o primeiro "ç"
# derruba o script com UnicodeEncodeError e a conferência morre depois de ter passado.
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

# ── O que se espera de Porto Ferreira ───────────────────────────────────────────────
#
# Números conferidos contra a usina real, não escolhidos: 20 inversores em 5 skids
# (`exports/raw/options`, 05/09/2026), 5 medidores de fronteira, estação que coleta POA e
# GHI e não coleta vento, e relé de temperatura. Se algum deles mudar, é porque a usina
# mudou (inversor comissionado, skid novo) ou porque o formato mudou — as duas merecem
# parar a conferência e ser olhadas, e é por isso que o esperado está escrito aqui em vez
# de ser calculado a partir da própria resposta.
USINA_PADRAO = 4  # Porto Ferreira — 7,4 MWp, a maior do escopo, o pior caso de medida
DIAS_PADRAO = 31  # o teto do passo de 5 min, que é o maior orçamento alcançável
PASSO_PADRAO = "5m"
BALDES_POR_DIA = {"5m": 288, "15m": 96, "1h": 24, "1d": 1}

#: Quantos skids entram no arquivo do diretor. Dois, e não um: com um só, uma soma que
#: ignorasse a seleção e devolvesse a usina inteira ainda daria UMA coluna, e a conferência
#: passaria sem ter provado nada.
SKIDS_ESCOLHIDOS = 2

# Tetos de dias por passo, do lado do meuWatt (`MAX_DAYS` em `src/exports/service.py`).
# Repetidos aqui só para MONTAR o pedido que tem de ser recusado — a palavra final é do
# servidor, e é justamente ela que a conferência quer ouvir.
TETO_5M = 31

#: O que NUNCA pode sair pela nossa porta. `wh_per_pulse`, `rtc` e `rtp` são parâmetros de
#: calibração do medidor — a constante de pulso e as relações de transformação com que se
#: converte pulso em kWh faturado. Quem baixa a planilha não os escolhe. O `slug` é
#: transporte para uma URL chamada com a credencial de serviço: o dia em que ele aparecer
#: numa resposta é o dia em que alguém descobre o nome que abre a porta do outro lado.
PROIBIDOS = ("wh_per_pulse", "rtc", "rtp")


class Falhou(Exception):
    """Uma conferência que não passou. Carrega a frase que o operador precisa ler."""


# ── Impressão ───────────────────────────────────────────────────────────────────────


class Placar:
    """Coleta as conferências.

    É objeto, e não uma lista global, porque o modo `--provar` precisa de um placar
    DESCARTÁVEL por mutação: contar as falhas de uma mutação no mesmo balde das falhas de
    verdade misturaria "o gate pegou o defeito que eu plantei" com "o gate achou um defeito
    real", que são notícias opostas.
    """

    def __init__(self) -> None:
        self.falhas: list[str] = []
        self.total = 0

    def __call__(self, nome: str, ok: bool, detalhe: str) -> bool:
        """Uma conferência nomeada. Imprime sempre o valor MEDIDO, passando ou não —
        "ok" sem número não deixa ninguém desconfiar quando o número muda."""
        self.total += 1
        print(f"  {'ok  ' if ok else 'FALHOU'} {nome}: {detalhe}")
        if not ok:
            self.falhas.append(f"{nome} — {detalhe}")
        return ok


Confere = Callable[[str, bool, str], bool]


def secao(titulo: str) -> None:
    print(f"\n{'=' * 78}\n{titulo}\n{'=' * 78}")


def mib(n: float) -> str:
    return f"{n / 1024 / 1024:.2f} MiB"


# ── Subir o BFF ─────────────────────────────────────────────────────────────────────


def _porta_livre() -> int:
    with socket.socket() as s:
        s.bind(("127.0.0.1", 0))
        return int(s.getsockname()[1])


def subir_bff() -> tuple[str, subprocess.Popen[bytes]]:
    """Um uvicorn próprio, em porta livre, com a saída engolida.

    Sem `--reload` e sem `--workers`, de propósito: assim o servidor roda NO processo que
    nasce aqui, e `proc.pid` é o worker — é o que torna a medida de memória possível. Com
    reload haveria um supervisor no meio e o RSS medido seria o do processo errado.

    A saída do servidor vai para `DEVNULL` porque ela atrapalha a leitura da conferência —
    mas o processo é morto no `finally` de quem chamou, sempre: um uvicorn órfão segurando
    porta é o tipo de sujeira que só aparece na terceira execução.
    """
    porta = _porta_livre()
    proc = subprocess.Popen(
        [sys.executable, "-m", "uvicorn", "app.main:app", "--port", str(porta), "--log-level", "warning"],
        cwd=str(RAIZ),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env={**os.environ, "PYTHONPATH": str(RAIZ)},
    )
    url = f"http://127.0.0.1:{porta}"
    for _ in range(120):
        if proc.poll() is not None:
            raise Falhou(f"o uvicorn morreu ao subir (código {proc.returncode}) — rode-o à mão para ver o erro")
        try:
            if httpx.get(f"{url}/health", timeout=1.0).status_code == 200:
                print(f"  BFF local no ar em {url}  (pid {proc.pid})")
                return url, proc
        except httpx.HTTPError:
            time.sleep(0.5)
    proc.kill()
    raise Falhou("o BFF local não respondeu /health em 60 s")


def token_do_gestor(user_id: int) -> str:
    """O JWT do cliente, emitido aqui — o script não tem a senha do dono, e não deve ter.

    Nunca impresso, nem em pedaço: o valor completo abre as sete usinas dele, e a saída
    deste script vai parar em log de terminal e em relato de agente.
    """
    from dotenv import load_dotenv

    load_dotenv(str(RAIZ / ".env"))
    from app.core.security import criar_token

    token, _ = criar_token(user_id)
    return token


# ── A memória do worker ─────────────────────────────────────────────────────────────


class Memoria:
    """Amostra o RSS do worker enquanto os bytes atravessam.

    Existe por uma pergunta que só se responde com número: `StreamingResponse` sobre
    `httpx.stream` é fluxo de verdade, ou é um buffer com outro nome? Se fosse buffer, o
    pico subiria o TAMANHO DO ARQUIVO (e mais uma cópia, no corpo do Starlette). Fluxo sobe
    o tamanho de alguns pedaços em trânsito. A diferença entre +2,4 MiB e +0,4 MiB por
    download é o que decide se dez clientes simultâneos cabem no contêiner.

    Amostra em thread, a cada 20 ms, porque o pico dura o tempo de um pedaço: medir só antes
    e depois não veria nada. Sem `psutil`, e no modo `--url` (o worker é de outra máquina),
    devolve `None` e quem chamou diz isso em voz alta em vez de inventar.
    """

    def __init__(self, pid: int | None) -> None:
        self.proc: Any = None
        self.motivo: str | None = None
        if pid is None:
            self.motivo = "o BFF é remoto (--url): o worker não está nesta máquina"
            return
        try:
            import psutil
        except ModuleNotFoundError:
            self.motivo = "sem `psutil` (pip install psutil)"
            return
        try:
            self.proc = psutil.Process(pid)
        except Exception as exc:  # noqa: BLE001
            self.motivo = f"não deu para observar o pid {pid}: {exc}"

    def rss(self) -> int | None:
        if self.proc is None:
            return None
        try:
            return int(self.proc.memory_info().rss)
        except Exception:  # noqa: BLE001
            return None

    def observar(self) -> "_Observacao":
        return _Observacao(self)


class _Observacao:
    """A janela de observação de um download."""

    def __init__(self, mem: Memoria) -> None:
        self.mem = mem
        self.base: int | None = None
        self.pico: int | None = None
        self._parar = threading.Event()
        self._t: threading.Thread | None = None

    def __enter__(self) -> "_Observacao":
        self.base = self.pico = self.mem.rss()
        if self.base is None:
            return self

        def laco() -> None:
            while not self._parar.wait(0.02):
                v = self.mem.rss()
                if v is not None and (self.pico is None or v > self.pico):
                    self.pico = v

        self._t = threading.Thread(target=laco, daemon=True)
        self._t.start()
        return self

    def __exit__(self, *_: Any) -> None:
        self._parar.set()
        if self._t is not None:
            self._t.join(timeout=2)


# ── As rotas ────────────────────────────────────────────────────────────────────────


def achar_rotas(url: str) -> tuple[str | None, str | None]:
    """Descobre no `/openapi.json` quais são, de fato, as rotas da exportação.

    Existe porque o modo de falhar mais provável deste script não é o arquivo vir errado —
    é a rota ter outro nome. Um 404 cru diria "não achei" e mandaria procurar no lugar
    errado; aqui a conferência consegue dizer *"a rota da exportação não está montada"*, que
    é uma frase acionável. É também a única pista honesta quando se roda contra produção: em
    05/09/2026 o `/openapi.json` de lá respondia 404, e é assim que se descobre que o build
    no ar é anterior a estas rotas — em vez de acusar a feature de estar quebrada.
    """
    try:
        doc = httpx.get(f"{url}/openapi.json", timeout=20.0).json()
    except (httpx.HTTPError, ValueError):
        return None, None
    opcoes = arquivo = None
    for caminho, verbos in doc.get("paths", {}).items():
        if "dados" not in caminho and "export" not in caminho:
            continue
        if "get" in verbos and opcoes is None and ("opcoes" in caminho or "options" in caminho):
            opcoes = caminho
        if "post" in verbos and arquivo is None and ("arquivo" in caminho or "raw" in caminho):
            arquivo = caminho
    return opcoes, arquivo


# ══════════════════════════════════════════════════════════════════════════════
# As conferências — funções PURAS sobre dados já colhidos
#
# Todas recebem os dados prontos e um `confere`. É essa forma que permite ao modo
# `--provar` alimentá-las com dados estragados e exigir que reprovem: a mutação
# exercita EXATAMENTE o mesmo código que a conferência de verdade, e não uma cópia
# que poderia divergir e continuar dizendo "ok" enquanto a de verdade quebrasse.
# ══════════════════════════════════════════════════════════════════════════════


def conferir_vazamento(bruto: str, slug: str, cf: Confere) -> None:
    """O que o meuWatt manda e a nossa porta não repassa.

    A resposta de lá traz, em cada leitor, `wh_per_pulse`, `rtc` e `rtp` — a constante de
    pulso e as relações de transformação com que se converte pulso em kWh faturado —, além
    do `slug` e do `name` da usina no sistema deles. Nada disso é escolha de quem baixa a
    planilha, e o slug é a chave que abre a mw-api com a credencial de serviço.

    A varredura é feita no TEXTO CRU da resposta e não no objeto já tipado, de propósito:
    um campo extra que escapasse pelo `model_extra` do Pydantic, ou um dicionário aberto
    como `estacao.colunas`, não apareceria numa checagem de atributos — apareceria aqui.
    """
    achados = [p for p in PROIBIDOS if p in bruto]
    cf(
        "os parâmetros de calibração do medidor não saem por esta porta",
        not achados,
        f"procurei {list(PROIBIDOS)} no corpo inteiro; achei {achados or 'nenhum'}",
    )
    # O slug é procurado pelo VALOR (`porto-ferreira`) e também pela CHAVE: uma resposta que
    # o publicasse com outro nome ("plant_key", "mw") passaria por uma busca só de chave, e
    # uma que publicasse o slug de OUTRA usina passaria por uma busca só de valor.
    cf(
        "o slug do meuWatt não sai por esta porta",
        slug not in bruto and '"slug"' not in bruto,
        f"procurei o valor do slug e a chave 'slug' no corpo inteiro ({len(bruto)} bytes)",
    )


@dataclass
class Aba:
    """Uma aba lida do XLSX, numa passada só."""

    nome: str
    linhas: int  # com o cabeçalho
    cabecalho: tuple[Any, ...]
    linha2: tuple[Any, ...] | None
    #: Soma dos valores numéricos de cada coluna, por rótulo do cabeçalho. É o que permite
    #: provar que a seleção parcial MUDOU A SOMA sem baixar um terceiro arquivo.
    somas: dict[str, float] = field(default_factory=dict)

    @property
    def colunas(self) -> int:
        return len(self.cabecalho)

    @property
    def dados(self) -> int:
        return max(0, self.linhas - 1)


@dataclass
class Planilha:
    abas: dict[str, Aba]

    @property
    def nomes(self) -> list[str]:
        return list(self.abas)


def abrir_planilha(conteudo: bytes) -> Planilha:
    """Abre o XLSX e mede cada aba numa passada só.

    `max_row`/`max_column` são `None` num arquivo escrito em modo write-only: o XLSX sai
    sem a dimensão declarada. Contar é a única leitura honesta — e é o que o Excel faz
    também.
    """
    try:
        import openpyxl
    except ModuleNotFoundError:
        raise Falhou(
            "sem `openpyxl` não dá para conferir o arquivo, só o tamanho dele — e tamanho "
            "não prova que abre. Rode: pip install openpyxl"
        ) from None

    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    try:
        abas: dict[str, Aba] = {}
        for nome in wb.sheetnames:
            ws = wb[nome]
            cab: tuple[Any, ...] = ()
            linha2: tuple[Any, ...] | None = None
            somas: dict[str, float] = {}
            n = 0
            for i, linha in enumerate(ws.iter_rows(values_only=True)):
                n += 1
                if i == 0:
                    cab = linha
                    continue
                if i == 1:
                    linha2 = linha
                for j, v in enumerate(linha):
                    # `bool` sai fora antes: em Python `True` soma como 1.0 sem reclamar, e
                    # uma coluna de status booleano viraria energia.
                    if j >= len(cab) or isinstance(v, bool) or not isinstance(v, (int, float)):
                        continue
                    rot = str(cab[j])
                    somas[rot] = somas.get(rot, 0.0) + float(v)
            abas[nome] = Aba(nome, n, cab, linha2, somas)
        return Planilha(abas)
    finally:
        wb.close()


def conferir_transporte(
    rotulo: str, conteudo: bytes, cabecalhos: httpx.Headers, segundos: float, cf: Confere
) -> None:
    """Os cabeçalhos e os bytes, antes de olhar o conteúdo."""
    cf(f"[{rotulo}] o download acabou dentro do prazo de leitura (120 s)", segundos < 120, f"{segundos:.1f}s")
    cf(
        f"[{rotulo}] veio como planilha, e não como JSON de erro",
        "spreadsheetml" in (cabecalhos.get("content-type") or ""),
        cabecalhos.get("content-type") or "(sem content-type)",
    )
    cf(
        f"[{rotulo}] veio como anexo, com nome de arquivo",
        "attachment" in (cabecalhos.get("content-disposition") or ""),
        (cabecalhos.get("content-disposition") or "(sem content-disposition)")[:130],
    )
    # Um XLSX é um ZIP. Se o corpo foi recodificado no caminho (um proxy que comprime e um
    # `Content-Length` do tamanho errado), o arquivo chega truncado e só o Excel percebe.
    cf(f"[{rotulo}] os bytes começam com a assinatura de um ZIP", conteudo[:2] == b"PK", repr(conteudo[:4]))
    # `Content-Length` que não bate com o corpo é o defeito exato que `pacotes.py` descreve:
    # o navegador corta o arquivo no número prometido, ou fica esperando bytes que não vêm.
    # Ausente é ACEITÁVEL (barra de progresso cega é o mal menor); errado, nunca.
    prometido = cabecalhos.get("content-length")
    cf(
        f"[{rotulo}] o tamanho prometido bate com o corpo recebido (ou não é prometido)",
        prometido is None or int(prometido) == len(conteudo),
        f"content-length={prometido} · recebidos {len(conteudo)}",
    )


def conferir_pesado(p: Planilha, baldes: int, n_inversores: int, cf: Confere) -> None:
    """O arquivo de todos os blocos: as seis abas, um instante por balde, uma coluna por
    inversor.

    O número de colunas não é "≥ alguma coisa": é a conta do contrato, escrita por extenso
    para poder ser conferida por quem lê. `sheet_inversores` (mw-api) emite, em `lista`, uma
    coluna por inversor para cada variável de valor, mais uma coluna de minutos desligados
    por inversor quando `paradas` foi pedida, mais os totais da usina — e o instante na
    frente. Um "≥ 21" passaria com metade dos inversores faltando.
    """
    esperadas = ["Leia-me", "Inversores", "Paradas", "Estação", "Fronteira", "Sistema"]
    cf("as seis abas do pedido completo, com estes nomes", p.nomes == esperadas, f"{p.nomes}")
    # A aba Leia-me é a razão de a TELA poder ser curta: unidade, fonte de cada coluna,
    # mapa das séries e avisos moram nela. Sem ela, "Energia (kWh)" numa coluna vira
    # adivinhação — e a tela teria de explicar tudo o que o arquivo já explica.
    leia = p.abas.get("Leia-me")
    cf(
        "o Leia-me não veio vazio",
        leia is not None and leia.dados >= 10,
        f"{leia.dados if leia else 0} linha(s) de conteúdo",
    )

    inv = p.abas.get("Inversores")
    if inv is None:
        cf("a aba Inversores existe", False, f"abas: {p.nomes}")
        return

    cf(
        "a aba Inversores tem um instante por balde do período",
        inv.dados == baldes,
        f"{inv.dados} linhas de dado (esperado {baldes} = dias × baldes/dia)",
    )
    # 2 variáveis de valor (geração, potência) × N inversores
    #   + 1 coluna de minutos desligados por inversor (a variável `paradas`)
    #   + 2 totais da usina (geração, potência) + 1 total de minutos desligados
    #   + a coluna do instante
    esperado_col = n_inversores * 3 + 3 + 1
    cf(
        "há uma coluna por inversor para cada variável, mais os totais da usina",
        inv.colunas == esperado_col,
        f"{inv.colunas} colunas (esperado {n_inversores}×3 + 3 + 1 = {esperado_col})",
    )
    cf(
        "a primeira coluna é o instante, em BRT",
        bool(inv.cabecalho) and "BRT" in str(inv.cabecalho[0]),
        repr(inv.cabecalho[0] if inv.cabecalho else None),
    )
    usina = [c for c in inv.cabecalho if str(c).startswith("Usina · ")]
    cf("os totais da usina fecham a aba", len(usina) == 3, f"{len(usina)}: {[str(c) for c in usina]}")
    # REGRA 0 dentro do arquivo: num passo sub-diário a linha 2 é meia-noite, e à
    # meia-noite o inversor não gera. As células vêm VAZIAS, não zero — quem somar a
    # coluna no Excel não vai contar madrugada como produção medida. Um dia em que isso
    # virar 0.0 é uma regressão silenciosa que só um olho treinado pegaria.
    celulas = (inv.linha2 or ())[1:]
    cf(
        "à meia-noite a ausência vem vazia, nunca zero",
        all(v is None for v in celulas) or not celulas,
        f"{sum(1 for v in celulas if v is not None)} célula(s) preenchidas na linha 2 de {len(celulas)}",
    )
    for aba in ("Paradas", "Estação", "Fronteira", "Sistema"):
        a = p.abas.get(aba)
        cf(f"a aba {aba} tem cabeçalho", a is not None and a.colunas > 0, f"{a.colunas if a else 0} coluna(s)")


def conferir_por_skid(p: Planilha, nomes_skids: list[str], baldes: int, cf: Confere) -> None:
    """O arquivo do diretor: a seleção parcial atravessou, e a soma mudou junto.

    ⚠ A aba **não se chama "Inversores"**: `sheet_inversores`, na mw-api, termina em
    `Sheet("Inversores" if agr == "lista" else "Skids", …)`. Quem agrupa por skid recebe uma
    aba chamada **Skids**. Está escrito aqui porque é o tipo de detalhe que uma tela repete
    errado ("baixe a aba Inversores") e ninguém confere.

    O que esta conferência guarda, e a do arquivo pesado não guarda: um BFF que ignorasse
    `series` — descartando a lista, ou mandando `null` no lugar dela — produziria um arquivo
    com os CINCO skids, que abre, tem as abas certas e passaria por qualquer checagem de
    formato. Aqui ele reprova em duas frentes: sobram colunas, e o total da usina deixa de
    ser a soma dos skids escolhidos.
    """
    cf(
        "agrupar por skid muda o nome da aba (Skids, não Inversores)",
        "Skids" in p.nomes and "Inversores" not in p.nomes,
        f"abas: {p.nomes}",
    )
    sk = p.abas.get("Skids")
    if sk is None:
        return

    cf(
        "a aba Skids tem um instante por balde do período",
        sk.dados == baldes,
        f"{sk.dados} linhas de dado (esperado {baldes})",
    )
    colunas_skid = [str(c) for c in sk.cabecalho if " · " in str(c) and not str(c).startswith("Usina · ")]
    cf(
        f"só os {len(nomes_skids)} skids escolhidos viraram coluna",
        len(colunas_skid) == len(nomes_skids),
        f"{len(colunas_skid)} coluna(s) de skid: {colunas_skid}",
    )
    rotulos = {c.split(" · ")[0] for c in colunas_skid}
    cf(
        "e são exatamente os skids que foram marcados",
        rotulos == set(nomes_skids),
        f"no arquivo {sorted(rotulos)} · pedidos {sorted(nomes_skids)}",
    )
    # A prova de que a seleção mudou a SOMA, e não só as colunas: o total da usina desta
    # planilha é a soma dos skids que estão nela. Se o BFF tivesse deixado a seleção cair,
    # o total seria o da usina inteira e não fecharia com as colunas presentes.
    col_total = next((str(c) for c in sk.cabecalho if str(c).startswith("Usina · ")), None)
    soma_skids = sum(sk.somas.get(c, 0.0) for c in colunas_skid)
    total = sk.somas.get(col_total or "", 0.0)
    # Tolerância de arredondamento: a mw-api arredonda cada célula com `_r(v, 3)` antes de
    # escrever, então o total nunca é bit a bit a soma das partes — é a soma dos
    # arredondados contra o arredondado da soma, e a diferença cresce com o nº de linhas.
    folga = max(1.0, abs(total) * 1e-6 + sk.dados * 0.002)
    cf(
        "o total da usina é a soma dos skids escolhidos, não a da usina inteira",
        col_total is not None and abs(total - soma_skids) <= folga,
        f"total={total:.3f} · soma dos skids={soma_skids:.3f} · "
        f"diferença={abs(total - soma_skids):.3f} (folga {folga:.3f})",
    )


def conferir_soma_mudou(pesado: Planilha, skid: Planilha, cf: Confere) -> None:
    """Os dois arquivos, lado a lado: o mesmo período, somas diferentes.

    É a conferência que fecha o pedido do dono. O arquivo pesado tem os 20 inversores; o do
    diretor tem os 8 de dois skids. Se os totais fossem iguais, a seleção não teria
    atravessado — e cada arquivo, olhado sozinho, pareceria perfeito.
    """
    a = pesado.abas.get("Inversores")
    b = skid.abas.get("Skids")
    if a is None or b is None:
        cf("os dois arquivos têm a aba de inversores", False, f"{pesado.nomes} × {skid.nomes}")
        return
    ca = next((str(c) for c in a.cabecalho if str(c).startswith("Usina · Energia")), None)
    cb = next((str(c) for c in b.cabecalho if str(c).startswith("Usina · Energia")), None)
    ta, tb = a.somas.get(ca or "", 0.0), b.somas.get(cb or "", 0.0)
    cf(
        "a usina inteira gerou mais que os dois skids escolhidos",
        ta > tb > 0,
        f"todos os inversores = {ta:,.0f} kWh · 2 skids = {tb:,.0f} kWh".replace(",", "."),
    )


def conferir_identidade_404(fora: httpx.Response, inexistente: httpx.Response, cf: Confere) -> None:
    """Usina que existe e não é sua, e usina que não existe: a MESMA resposta.

    Responder "proibido" para a primeira e "não encontrada" para a segunda transformaria a
    rota num oráculo: um cliente varreria os ids e saberia exatamente quantas usinas o
    sistema tem e quais são. A propriedade que importa não é o código nem a frase — é que as
    duas sejam **indistinguíveis**, e é por isso que a comparação é byte a byte no corpo cru,
    e não `assert status == 404` duas vezes.
    """
    cf(
        "usina de outro cliente e usina inexistente respondem o mesmo código",
        fora.status_code == inexistente.status_code == 404,
        f"fora do escopo {fora.status_code} · inexistente {inexistente.status_code}",
    )
    cf(
        "e o mesmo corpo, byte a byte",
        fora.content == inexistente.content,
        f"{fora.content[:90]!r} × {inexistente.content[:90]!r}",
    )


# ══════════════════════════════════════════════════════════════════════════════
# A colheita — o que fala com a rede
# ══════════════════════════════════════════════════════════════════════════════


@dataclass
class Baixado:
    conteudo: bytes
    cabecalhos: httpx.Headers
    cabecalho_em: float
    corpo_em: float
    rss_base: int | None
    rss_pico: int | None

    @property
    def total(self) -> float:
        return self.cabecalho_em + self.corpo_em


def baixar(c: httpx.Client, rota: str, usina: int, pedido: dict[str, Any], mem: Memoria) -> Baixado:
    """Um POST de arquivo, medido em duas partes.

    A separação entre cabeçalho e corpo não é preciosismo: ela é a prova de ONDE a espera
    mora. A mw-api gera o XLSX inteiro (`to_thread(write_xlsx)`) antes de responder e só
    então manda um `FileResponse` de um temporário — então o cabeçalho demora dezenas de
    segundos e o corpo passa em pouco mais de um. É por isso que a tela não pode ter barra de
    progresso com porcentagem: não existe progresso para medir, existe espera.
    """
    with mem.observar() as obs:
        t0 = time.perf_counter()
        with c.stream("POST", rota, params={"usina_id": usina}, json=pedido) as r:
            cabecalho_em = time.perf_counter() - t0
            if r.status_code != 200:
                r.read()
                raise Falhou(f"o arquivo veio {r.status_code} em {cabecalho_em:.1f}s: {r.text[:500]}")
            t1 = time.perf_counter()
            conteudo = r.read()
            corpo_em = time.perf_counter() - t1
            cabecalhos = r.headers
    return Baixado(conteudo, cabecalhos, cabecalho_em, corpo_em, obs.base, obs.pico)


def relatar(rotulo: str, b: Baixado, mem: Memoria) -> None:
    print(f"  [{rotulo}] cabeçalho em ... {b.cabecalho_em:.1f}s  (a mw-api gera o arquivo INTEIRO antes de responder)")
    print(f"  [{rotulo}] corpo em ....... {b.corpo_em:.1f}s")
    print(f"  [{rotulo}] total .......... {b.total:.1f}s")
    print(f"  [{rotulo}] tamanho ........ {len(b.conteudo)} bytes ({mib(len(b.conteudo))})")
    if b.rss_base is not None and b.rss_pico is not None:
        d = b.rss_pico - b.rss_base
        pct = 100.0 * d / max(1, len(b.conteudo))
        print(
            f"  [{rotulo}] memória ........ base {mib(b.rss_base)} → pico {mib(b.rss_pico)} "
            f"(+{d / 1024:.0f} KiB = {pct:.0f}% do arquivo)"
        )
    else:
        print(f"  [{rotulo}] memória ........ não medida: {mem.motivo}")


def colher_opcoes(c: httpx.Client, rota: str, usina: int, cf: Confere) -> tuple[dict[str, Any], str]:
    secao(f"1. O que esta usina pode oferecer  ·  GET {rota}")
    t0 = time.perf_counter()
    r = c.get(rota, params={"usina_id": usina})
    dt = time.perf_counter() - t0
    if r.status_code != 200:
        raise Falhou(f"as opções vieram {r.status_code}: {r.text[:400]}")
    bruto = r.text
    o = r.json()
    print(f"  {dt:.1f}s  ·  {len(bruto)} bytes")

    skids = o.get("skids") or []
    series = sum(len(s.get("series") or []) for s in skids)
    est = o.get("estacao") or {}
    colunas = [k for k, v in (est.get("colunas") or {}).items() if v]
    leitores = o.get("leitores") or []
    ret = o.get("retencao") or {}
    lim = o.get("limites") or {}
    usina_out = o.get("usina") or {}

    print(f"  usina .......... {usina_out.get('nome')} · {usina_out.get('capacidade_kwp')} kWp")
    print(f"  inversores ..... {series} em {len(skids)} skid(s): {[s.get('nome') for s in skids]}")
    print(f"  estação ........ disponível={est.get('disponivel')} · colunas com dado: {', '.join(colunas) or '—'}")
    print(f"  fronteira ...... {len(leitores)} leitor(es)")
    print(
        f"  sistema ........ pr={(o.get('sistema') or {}).get('pr')} "
        f"produtividade={(o.get('sistema') or {}).get('produtividade')}"
    )
    print(f"  retenção ....... snapshots desde {ret.get('snapshots_desde')} · SSU desde {ret.get('ssu_desde')}")
    print(f"  tetos (dias) ... {json.dumps(lim, ensure_ascii=False)}")

    cf(
        "a usina sai com o id e o nome DESTE sistema",
        usina_out.get("id") == usina and bool(usina_out.get("nome")),
        f"id={usina_out.get('id')} nome={usina_out.get('nome')!r}",
    )
    cf("as opções trazem inversores", series > 0, f"{series} séries")
    cf(
        "as opções trazem os tetos por passo",
        all(k in lim for k in ("native", "5m", "15m", "1h", "1d")),
        f"chaves: {sorted(lim)}",
    )
    cf(
        "as opções dizem desde quando existe o acervo",
        bool(ret.get("snapshots_desde") and ret.get("ssu_desde")),
        f"{ret.get('snapshots_desde')} / {ret.get('ssu_desde')}",
    )
    # A ausência tem de vir dita, não sumida: a estação de Porto Ferreira mede POA e GHI e
    # NÃO mede vento. Se `colunas` viesse tudo `true`, a tela ofereceria coluna vazia.
    cf(
        "a estação diz coluna a coluna o que coleta",
        isinstance(est.get("colunas"), dict) and len(est["colunas"]) > 0,
        f"{len(est.get('colunas') or {})} colunas declaradas, {len(colunas)} com dado",
    )
    return o, bruto


def colher_escopo(
    url: str,
    rota: str,
    tokens: dict[int, str],
    dono: int,
    forasteiro: int,
    usina_ok: int,
    usina_sem_mw: int,
    cf: Confere,
) -> None:
    """As quatro sondas de escopo, pela rota de verdade.

    ⚠ **Quem pergunta importa.** "Fora do escopo" não é uma usina especial: é a MESMA usina,
    perguntada por outra pessoa. No banco de hoje o dono enxerga todas as usinas que existem
    (7 de 7), então a única forma honesta de exercitar esse caminho é usar um segundo
    usuário — e é o que se faz aqui, dizendo qual token pergunta o quê. Inventar um id "fora
    do escopo" para o dono seria montar um cenário que não é o do sistema, e o verde não
    valeria nada.
    """
    secao("2. O escopo  ·  quem pode ver o quê")
    prazo = httpx.Timeout(connect=5.0, read=60.0, write=30.0, pool=5.0)
    respostas: dict[str, httpx.Response] = {}
    # (rótulo, usina, quem pergunta, o que tem de responder)
    plano = [
        ("no escopo", usina_ok, dono, 200),
        ("sem monitoramento", usina_sem_mw, dono, 404),
        ("de outro cliente", usina_ok, forasteiro, 404),
        ("inexistente", 9999, forasteiro, 404),
    ]
    for nome, uid, quem, esperado in plano:
        with httpx.Client(base_url=url, headers={"Authorization": f"Bearer {tokens[quem]}"}, timeout=prazo) as c:
            r = c.get(rota, params={"usina_id": uid})
        respostas[nome] = r
        corpo = r.text[:110] if "json" in (r.headers.get("content-type") or "") else f"({len(r.content)} bytes)"
        print(f"  usina_id={uid:<5} usuário {quem}  {nome:<20} → {r.status_code}  {corpo}")
        cf(f"usina {nome} responde {esperado}", r.status_code == esperado, f"HTTP {r.status_code}")

    sem = respostas["sem monitoramento"]
    cf(
        "a usina sem monitoramento diz isso, e não 'não encontrada'",
        "monitoramento" in sem.text.lower(),
        sem.text[:140],
    )
    conferir_identidade_404(respostas["de outro cliente"], respostas["inexistente"], cf)


def conferir_recusa(c: httpx.Client, rota: str, usina: int, cf: Confere) -> None:
    """Pedir 92 dias a 5 minutos tem de voltar recusado, DEPRESSA.

    É o limite que a tela promete impedir antes de deixar clicar. Aqui se prova a outra
    metade: quando um cliente velho, ou alguém montando o pedido na mão, fura a tela, o
    servidor recusa — e recusa em segundos, sem gastar os 35 s de gerar um arquivo que ele
    já sabe que não vai entregar. E o `motivo` tem de ATRAVESSAR o BFF: é o código estável
    que a tela traduz. Se o BFF engolir o motivo e devolver só "não deu para baixar", a tela
    perde a única frase útil que existia.
    """
    secao("6. A recusa  ·  92 dias a cada 5 minutos (o teto é 31)")
    pedido = {
        "inicio": "2026-06-01",
        "fim": "2026-08-31",
        "hora_inicio": "00:00",
        "hora_fim": "23:59",
        "passo": "5m",
        "inversores": {"variaveis": ["geracao"], "agrupamento": "lista"},
    }
    t0 = time.perf_counter()
    r = c.post(rota, params={"usina_id": usina}, json=pedido)
    dt = time.perf_counter() - t0
    print(f"  {r.status_code} em {dt:.1f}s")
    # O corpo só é impresso quando é texto. Um 200 aqui significaria que o pedido passou —
    # e aí o corpo é um XLSX: despejar `r.text` de um ZIP enche o terminal de lixo binário
    # bem na hora em que alguém está tentando ler por que a conferência falhou.
    tipo = (r.headers.get("content-type") or "").lower()
    print(f"  corpo: {r.text[:400] if 'json' in tipo or 'text' in tipo else f'({len(r.content)} bytes de {tipo!r})'}")

    cf("o pedido acima do teto é recusado", r.status_code == 400, f"HTTP {r.status_code}")
    cf(
        "a recusa é imediata, não depois de gerar",
        dt < 5.0,
        f"{dt:.1f}s (gerar o arquivo levaria dezenas de segundos)",
    )
    corpo: dict[str, Any] = {}
    try:
        lido = r.json()
        corpo = lido if isinstance(lido, dict) else {}
    except ValueError:
        pass
    # O `motivo` no PRIMEIRO nível, e não embrulhado em `detail`: é ele que a tela lê para
    # escolher entre `Erro` (com "Tentar de novo") e `Aviso` (sem). Um `muito_grande`
    # repetido dá exatamente o mesmo resultado — oferecer o botão seria crueldade.
    cf(
        "o motivo atravessa o BFF, no primeiro nível, para a tela traduzir",
        corpo.get("motivo") == "passo_excede_limite",
        f"motivo={corpo.get('motivo')!r} (esperado 'passo_excede_limite')",
    )
    # O `message` do meuWatt fala em balde, snapshots e SSU: foi escrito para o operador de
    # lá. Se ele vazar para cá, o cliente lê o vocabulário de um sistema que ele não abre.
    detalhe = str(corpo.get("detail") or "")
    cf(
        "o texto do upstream não é ecoado ao cliente",
        "aceita até" not in detalhe and "snapshot" not in detalhe.lower(),
        f"detail={detalhe!r}",
    )
    print(f"\n  (o teto de 5 min é {TETO_5M} dias; o pedido tinha 92)")


# ══════════════════════════════════════════════════════════════════════════════
# `--provar`: a mutação
# ══════════════════════════════════════════════════════════════════════════════


def _planilha_falsa(abas: dict[str, tuple[list[Any], list[list[Any]]]]) -> Planilha:
    """Uma planilha montada na memória, para o modo `--provar`.

    Passa pelo `abrir_planilha` de verdade — escreve um XLSX com `openpyxl`, serializa e lê
    de volta — em vez de fabricar os objetos `Aba` na mão. Se fabricasse, a mutação provaria
    que a conferência reprova um OBJETO errado, e não que ela reprova um ARQUIVO errado, que
    é a única coisa que interessa.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for nome, (cab, linhas) in abas.items():
        ws = wb.create_sheet(nome)
        ws.append(cab)
        for linha in linhas:
            ws.append(linha)
    buf = io.BytesIO()
    wb.save(buf)
    return abrir_planilha(buf.getvalue())


def _resposta_falsa(corpo: dict[str, Any], status: int = 404) -> httpx.Response:
    return httpx.Response(status, json=corpo, request=httpx.Request("GET", "http://x/"))


def provar_que_reprova() -> int:
    """Alimenta cada conferência com o defeito que ela existe para pegar.

    Um gate que nunca se viu reprovar é decoração: ele passa verde tanto quando o sistema
    está certo quanto quando a própria checagem está morta (um `and` que virou `or`, um campo
    renomeado que faz a busca não achar nada e concluir "não vazou"). Aqui cada conferência
    recebe o dado estragado e é OBRIGADA a acusar — e o modo sai com 1 se alguma delas
    deixar passar.

    Não fala com a rede: é o único modo que roda sem credencial e sem o meuWatt de pé.
    """
    secao("PROVA — cada conferência tem de REPROVAR o defeito que ela guarda")
    mutacoes: list[tuple[str, Callable[[Confere], None]]] = []

    # 1. O vazamento: a resposta volta a carregar o que a nossa porta tirou — a calibração
    #    do medidor e o nome que abre a mw-api com a credencial de serviço.
    bruto_ruim = json.dumps(
        {
            "usina": {"id": 4, "nome": "Porto Ferreira", "slug": "porto-ferreira"},
            "leitores": [{"id": 14, "nome": "Leitor 1", "wh_per_pulse": 0.3, "rtc": 20.0, "rtp": 70.0}],
        }
    )
    mutacoes.append(
        (
            "o slug e a calibração do medidor voltam a sair pela porta",
            lambda cf: conferir_vazamento(bruto_ruim, "porto-ferreira", cf),
        )
    )

    # 2. O arquivo pesado perde a aba Leia-me (e com ela as unidades e o mapa das séries) e
    #    volta com metade dos inversores — o arquivo abre, e está errado.
    cab_meio = ["Início (BRT)"] + [f"Inv {i} · Energia (kWh)" for i in range(1, 11)]
    mutacoes.append(
        (
            "o arquivo pesado perde a aba Leia-me e metade dos inversores",
            lambda cf: conferir_pesado(
                _planilha_falsa({"Inversores": (cab_meio, [[None] * len(cab_meio) for _ in range(3)])}),
                baldes=8928,
                n_inversores=20,
                cf=cf,
            ),
        )
    )

    # 3. A meia-noite preenchida com zero: a regressão silenciosa da REGRA 0 dentro do
    #    arquivo. O Excel somaria madrugada como produção medida.
    cab_cheio = (
        ["Início (BRT)"]
        + [f"Inv {i} · {v}" for i in range(1, 21) for v in ("Energia (kWh)", "Potência média (kW)", "Min. desligado")]
        + ["Usina · Energia (kWh)", "Usina · Potência média (kW)", "Usina · Min. desligado (soma dos inversores)"]
    )
    seis = {
        "Leia-me": (["campo", "valor"], [["x", i] for i in range(12)]),
        "Inversores": (cab_cheio, [["2026-08-05 00:00"] + [0.0] * (len(cab_cheio) - 1)]),
        "Paradas": (["Início (BRT)"], []),
        "Estação": (["Início (BRT)"], []),
        "Fronteira": (["Início (BRT)"], []),
        "Sistema": (["Início (BRT)"], []),
    }
    mutacoes.append(
        (
            "a meia-noite passa a vir com zero em vez de vazia",
            lambda cf: conferir_pesado(_planilha_falsa(seis), baldes=1, n_inversores=20, cf=cf),
        )
    )

    # 4. A seleção por skid cai no caminho: o arquivo volta com os cinco skids, e o total
    #    da usina deixa de fechar com as colunas presentes.
    cab_skid = ["Início (BRT)"] + [f"SKID-0{i} · Energia (kWh)" for i in range(1, 6)] + ["Usina · Energia (kWh)"]
    mutacoes.append(
        (
            "a seleção de skids cai e o arquivo volta com a usina inteira",
            lambda cf: conferir_por_skid(
                _planilha_falsa({"Skids": (cab_skid, [["2026-08-05 12:00", 10.0, 10.0, 10.0, 10.0, 10.0, 50.0]])}),
                ["SKID-01", "SKID-02"],
                baldes=1,
                cf=cf,
            ),
        )
    )

    # 5. Os dois 404 deixam de ser indistinguíveis: a rota vira um oráculo de ids.
    mutacoes.append(
        (
            "usina alheia e usina inexistente passam a responder frases diferentes",
            lambda cf: conferir_identidade_404(
                _resposta_falsa({"detail": "Acesso negado a esta usina."}, 403),
                _resposta_falsa({"detail": "Usina não encontrada."}),
                cf,
            ),
        )
    )

    # 6. Os dois arquivos passam a somar igual: sinal de que a seleção não mudou nada.
    mutacoes.append(
        (
            "os dois arquivos passam a somar a mesma energia",
            lambda cf: conferir_soma_mudou(
                _planilha_falsa({"Inversores": (["Início (BRT)", "Usina · Energia (kWh)"], [["x", 100.0]])}),
                _planilha_falsa({"Skids": (["Início (BRT)", "Usina · Energia (kWh)"], [["x", 100.0]])}),
                cf,
            ),
        )
    )

    cego = False
    for nome, roda in mutacoes:
        print(f"\n  MUTAÇÃO · {nome}")
        p = Placar()
        roda(p)
        if p.falhas:
            print(f"    → a conferência REPROVOU, como devia ({len(p.falhas)} de {p.total} checagens acusaram)")
        else:
            print(f"    → ⚠ a conferência DEIXOU PASSAR ({p.total} checagens, nenhuma acusou) — o gate está cego aqui")
            cego = True

    secao("PROVA — resultado")
    if cego:
        print("  ⚠ ao menos uma conferência não pegou o defeito que ela existe para pegar.")
        return 1
    print(f"  as {len(mutacoes)} mutações foram todas acusadas — as conferências têm dentes.")
    return 0


# ── Ferramentas do pedido ───────────────────────────────────────────────────────────


def _variaveis_de_estacao(o: dict[str, Any]) -> list[str]:
    """As variáveis de estação que ESTA usina coleta.

    Lidas das opções e não fixas no código: pedir `vento` a uma usina que não mede vento é
    um 400 (`bloco_indisponivel`) — o servidor recusa o bloco inteiro, e a conferência
    morreria acusando a feature por um defeito do pedido. É a mesma leitura que a tela faz
    para desabilitar a linha dizendo o motivo.

    `umidade` fica de fora sempre: ela vem no dicionário de colunas (medido: `false`) mas
    não é variável exportável do contrato — nenhuma estação a envia. A tela mostra a linha
    permanentemente desabilitada, com o motivo; aqui ela simplesmente não é pedida.
    """
    est = o.get("estacao") or {}
    vars_ = [k for k, v in (est.get("colunas") or {}).items() if v and k != "umidade"]
    if est.get("temp_ambiente_rele"):
        vars_.append("temp_ambiente_rele")
    return vars_ or ["poa"]


def _slug_da_usina(usina_id: int) -> str:
    """O slug do meuWatt desta usina, lido do NOSSO banco — o mesmo valor que a resposta não
    pode conter.

    Sem ele, a varredura de vazamento só saberia procurar a chave `slug`, e uma resposta que
    o publicasse com outro nome passaria. O valor de recuo é impossível de casar de propósito
    (`\\x00`): uma usina sem vínculo não pode fazer a varredura achar o slug vazio em toda
    resposta e reprovar tudo.
    """
    from dotenv import load_dotenv

    load_dotenv(str(RAIZ / ".env"))
    from app.core.db import SessionLocal
    from app.models import PlantLink

    db = SessionLocal()
    try:
        link = db.query(PlantLink).filter(PlantLink.id == usina_id).first()
        return (getattr(link, "mw_plant_slug", "") or "").strip() or "\x00sem-slug\x00"
    finally:
        db.close()


# ── Comando ─────────────────────────────────────────────────────────────────────────


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--url", help="BFF já no ar; sem isto, sobe um local")
    p.add_argument("--usina", type=int, default=USINA_PADRAO, help="id do vínculo (PlantLink)")
    p.add_argument("--dias", type=int, default=DIAS_PADRAO, help="tamanho do período, terminando ontem")
    # `native` fica de fora: ele não tem um número fixo de baldes por dia (é uma linha por
    # leitura que o equipamento mandou), e sem esse número a conferência de linhas não teria
    # o que comparar. Um passo que o script não sabe conferir é pior do que um passo que ele
    # não oferece — passaria verde sem ter verificado nada.
    p.add_argument(
        "--passo",
        default=PASSO_PADRAO,
        choices=sorted(BALDES_POR_DIA),
        help="passo do arquivo (`native` fica de fora: não tem baldes/dia fixo para conferir)",
    )
    p.add_argument("--usuario", type=int, default=2, help="id do usuário do Gestão Solar (o dono)")
    p.add_argument("--forasteiro", type=int, default=1, help="id de um usuário que NÃO enxerga a usina do teste")
    p.add_argument("--sem-monitoramento", type=int, default=8, help="usina do escopo sem vínculo com o meuWatt")
    p.add_argument(
        "--guardar",
        action="store_true",
        help="grava os .xlsx no temporário do sistema, para abrir no Excel e olhar",
    )
    p.add_argument(
        "--provar",
        action="store_true",
        help="não fala com ninguém: alimenta as conferências com defeitos e exige que reprovem",
    )
    args = p.parse_args()

    if args.provar:
        return provar_que_reprova()

    cf = Placar()
    proc = None
    t0 = time.perf_counter()
    try:
        if args.url:
            url = args.url.rstrip("/")
            print(f"  BFF: {url}")
        else:
            url, proc = subir_bff()

        opcoes, arquivo = achar_rotas(url)
        if not (opcoes and arquivo):
            raise Falhou(
                "a rota da exportação não está montada neste BFF — o `/openapi.json` não tem "
                f"GET de opções nem POST de arquivo sob /dados (achei opcoes={opcoes!r}, "
                f"arquivo={arquivo!r}). Não há o que conferir: a tela não teria de onde ler. "
                "Contra produção, é assim que se descobre que o build no ar é anterior a estas rotas."
            )
        print(f"  rotas: GET {opcoes} · POST {arquivo}")

        tokens = {
            args.usuario: token_do_gestor(args.usuario),
            args.forasteiro: token_do_gestor(args.forasteiro),
        }
        mem = Memoria(proc.pid if proc is not None else None)

        # O período termina ONTEM: o dia de hoje ainda está sendo escrito, e um arquivo que
        # inclui o dia corrente muda de conteúdo entre duas execuções do mesmo comando —
        # exatamente o que uma conferência não pode ter.
        fim = date.today() - timedelta(days=1)
        inicio = fim - timedelta(days=args.dias - 1)
        baldes = args.dias * BALDES_POR_DIA[args.passo]

        prazo = httpx.Timeout(connect=5.0, read=180.0, write=30.0, pool=5.0)
        with httpx.Client(
            base_url=url, headers={"Authorization": f"Bearer {tokens[args.usuario]}"}, timeout=prazo
        ) as c:
            o, bruto = colher_opcoes(c, opcoes, args.usina, cf)
            conferir_vazamento(bruto, _slug_da_usina(args.usina), cf)

            colher_escopo(url, opcoes, tokens, args.usuario, args.forasteiro, args.usina, args.sem_monitoramento, cf)

            skids = o.get("skids") or []
            n_inv = sum(len(s.get("series") or []) for s in skids)
            escolhidos = skids[:SKIDS_ESCOLHIDOS]
            nomes_skids = [str(s.get("nome")) for s in escolhidos]
            chaves = [str(x["chave"]) for s in escolhidos for x in (s.get("series") or [])]

            secao(
                f"3. Os dois arquivos  ·  POST {arquivo}  ·  {inicio}..{fim} "
                f"({args.dias} dias) a cada {args.passo}"
            )
            print(f"  esperado: {baldes} instantes ({args.dias} × {BALDES_POR_DIA[args.passo]} baldes/dia)")

            # O pesado: todos os blocos, todos os inversores. `series` fica FORA de propósito
            # — ausente quer dizer "todas, inclusive a que for comissionada no meio do
            # período", que é diferente de listar as de hoje.
            pedido_pesado: dict[str, Any] = {
                "inicio": inicio.isoformat(),
                "fim": fim.isoformat(),
                "hora_inicio": "00:00",
                "hora_fim": "23:59",
                "passo": args.passo,
                "inversores": {"variaveis": ["geracao", "potencia", "paradas"], "agrupamento": "lista"},
                "estacao": {"variaveis": _variaveis_de_estacao(o)},
                "fronteira": {"variaveis": ["energia"], "agrupamento": "leitor"},
                "sistema": {"variaveis": ["pr", "produtividade"], "agrupamento": "usina"},
            }
            print(f"\n  pedido PESADO: os quatro blocos, {n_inv} inversores em lista")
            print(f"                 estação: {pedido_pesado['estacao']['variaveis']}")
            pesado = baixar(c, arquivo, args.usina, pedido_pesado, mem)
            relatar("pesado", pesado, mem)

            # O do diretor: só inversores, agrupados por skid, com uma PARTE dos skids.
            pedido_skid: dict[str, Any] = {
                "inicio": inicio.isoformat(),
                "fim": fim.isoformat(),
                "hora_inicio": "00:00",
                "hora_fim": "23:59",
                "passo": args.passo,
                "inversores": {"variaveis": ["geracao"], "agrupamento": "skid", "series": chaves},
            }
            print(
                f"\n  pedido POR SKID: {len(escolhidos)} de {len(skids)} skids "
                f"({', '.join(nomes_skids)}) = {len(chaves)} de {n_inv} inversores"
            )
            porskid = baixar(c, arquivo, args.usina, pedido_skid, mem)
            relatar("por skid", porskid, mem)

            secao("4. O transporte  ·  os cabeçalhos que o navegador vai ler")
            conferir_transporte("pesado", pesado.conteudo, pesado.cabecalhos, pesado.total, cf)
            conferir_transporte("por skid", porskid.conteudo, porskid.cabecalhos, porskid.total, cf)

            secao("5. As planilhas, abertas")
            pl_pesado = abrir_planilha(pesado.conteudo)
            pl_skid = abrir_planilha(porskid.conteudo)
            for rot, pl in (("pesado", pl_pesado), ("por skid", pl_skid)):
                print(f"\n  [{rot}] abas: {pl.nomes}")
                for a in pl.abas.values():
                    print(f"    {a.nome:<12} {a.dados:>6} linha(s) de dado × {a.colunas:>3} coluna(s)")

            print("\n  — o arquivo pesado —")
            conferir_pesado(pl_pesado, baldes, n_inv, cf)
            inv = pl_pesado.abas.get("Inversores")
            if inv:
                print(f"    cabeçalho, 3 primeiras ... {inv.cabecalho[:3]}")
                print(f"    cabeçalho, 3 últimas ..... {inv.cabecalho[-3:]}")

            print("\n  — o arquivo do diretor (por skid) —")
            conferir_por_skid(pl_skid, nomes_skids, baldes, cf)
            sk = pl_skid.abas.get("Skids")
            if sk:
                print(f"    cabeçalho ................ {[str(x) for x in sk.cabecalho]}")

            print("\n  — os dois, lado a lado —")
            conferir_soma_mudou(pl_pesado, pl_skid, cf)

            if args.guardar:
                # Fora do repositório de propósito: `--guardar` existe para alguém ABRIR o
                # arquivo no Excel e olhar, não para versionar um binário de megabytes junto
                # do código. Ao lado do script, o primeiro `git add` distraído o levaria.
                tmp = Path(tempfile.gettempdir())
                for nome, dados in (("pesado", pesado.conteudo), ("por-skid", porskid.conteudo)):
                    destino = tmp / f"dados-{args.usina}-{inicio}_{fim}-{args.passo}-{nome}.xlsx"
                    destino.write_bytes(dados)
                    print(f"\n  gravado em ..... {destino}")

            conferir_recusa(c, arquivo, args.usina, cf)

    except Falhou as e:
        secao("PAROU")
        print(f"  {e}")
        return 1
    except httpx.HTTPError as e:
        secao("PAROU — transporte")
        print(f"  {type(e).__name__}: {e}")
        return 1
    finally:
        if proc is not None:
            proc.kill()
            proc.wait(timeout=10)

    secao(f"RESULTADO  ·  {time.perf_counter() - t0:.1f}s no total  ·  {cf.total} conferências")
    if cf.falhas:
        for f in cf.falhas:
            print(f"  FALHOU  {f}")
        print(f"\n  {len(cf.falhas)} conferência(s) não passaram.")
        return 1
    print("  todas passaram — o cliente baixa as duas planilhas, elas abrem, e a seleção por")
    print("  skid chega do outro lado mudando a soma.")
    print("\n  Para ver o gate reprovar: python scripts/conferir_exportacao.py --provar")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
